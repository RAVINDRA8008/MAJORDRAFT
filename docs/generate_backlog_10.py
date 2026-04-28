"""Generate AMERS backlog — 10 sprints, 20 user stories, reflecting actual implementation."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AMERS Backlog"

ASSIGNED = "RAVINDRA  JONNAGADLA  (RA2211028010187);NIKETH VATTUMILLI (RA2211028010144)"
CREATED_BY = "RAVINDRA  JONNAGADLA  (RA2211028010187)"
CREATED_DATE = "01/05/2026"

HEADERS = [
    "Task ID", "Task Name", "Bucket Name", "Progress", "Priority",
    "Assigned To", "Created By", "Created Date", "Start date", "Due date",
    "Is Recurring", "Late", "Completed Date", "Checklist Items", "Labels", "Description"
]

rows = [
    # ═══════════════════════════════════════════════════════════════
    # SPRINT 1 — Dataset & Pipeline Setup (UNCHANGED)
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "QLgcTO7TbkCIrJ_68jXKSskAIjHK",
        "name": "SPRINT 1 — Dataset & Pipeline Setup",
        "bucket": "IN PROGRESS",
        "progress": "Not started",
        "start": "12/07/2025", "due": "12/26/2025",
        "late": "true",
        "checklist": "Select DEAP dataset;Select IEMOCAP dataset;Document channels and annotations;Verify academic license;Define unified emotion labels",
        "labels": "Sprint 1;User Story 1",
        "desc": """User Story 1 — Dataset Selection, Documentation & Label Alignment
User Story
As a researcher, I want to select and document reliable EEG and speech datasets
so that the experimental foundation of the system is valid and reproducible.
Description
This user story covers dataset selection, documentation, and unified emotion label alignment for EEG and speech modalities.
Outcome
Datasets are selected, documented, and aligned under a common emotion label space.
Acceptance Criteria
EEG dataset is selected and documented
Speech dataset is selected and documented
Unified emotion labels are defined
Labels are consistently mapped across datasets
Documentation is reproducible"""
    },
    {
        "id": "oxyj9x-Zd068u-utm2Cs18kAPlyT",
        "name": "SPRINT 1 — Dataset & Pipeline Setup",
        "bucket": "IN PROGRESS",
        "progress": "Not started",
        "start": "12/17/2025", "due": "12/26/2025",
        "late": "true",
        "checklist": "Apply LOSO protocol;Define unified emotion labels;Fix random seeds;Run pilot preprocessing;Verify output shapes",
        "labels": "Sprint 1;User Story 2",
        "desc": """User Story 2 — Leakage-Free Data Splitting & Pipeline Validation
User Story
As a researcher, I want to create leakage-free data splits and validate the pipeline
so that evaluation results remain unbiased and reliable.
Description
This user story ensures subject-independent dataset splits and validates the preprocessing pipeline through pilot runs.
Outcome
Leakage-free dataset splits and a verified preprocessing pipeline.
Acceptance Criteria
LOSO protocol is implemented
No subject or speaker overlap exists
Pipeline runs without runtime errors
Output data shapes are correct"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 2 — EEG Feature Engineering & Baseline Modeling (UNCHANGED)
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "of1VQ25zLkKCxHjaIiNsJMkAHq1B",
        "name": "SPRINT 2 — EEG Feature Engineering & Baseline Modeling",
        "bucket": "UP NEXT",
        "progress": "Not started",
        "start": "12/27/2025", "due": "01/08/2026",
        "late": "true",
        "checklist": "Segment EEG into epochs;Apply band-pass filtering;Perform artifact suppression;Extract EEG features",
        "labels": "Sprint 2;User Story 3",
        "desc": """User Story 3 — EEG Preprocessing & Feature Extraction
User Story
As a system, I want to preprocess EEG signals and extract meaningful features
so that emotional patterns can be effectively learned by machine learning models.
Description
This user story covers EEG signal preprocessing, noise removal, epoching, and feature extraction required for emotion classification.
Outcome
Clean EEG data and structured feature vectors ready for model training.
Acceptance Criteria
EEG signals are filtered and cleaned
Artifacts are reduced or removed
EEG signals are segmented into epochs
Feature vectors are generated consistently"""
    },
    {
        "id": "E7Y5dKCweUKdCG2uT-RR8ckANzf6",
        "name": "SPRINT 2 — EEG Feature Engineering & Baseline Modeling",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "01/09/2026", "due": "01/21/2026",
        "late": "true",
        "checklist": "Implement classifier model;Train model without errors;Compute baseline accuracy and F1-score;Log and save results",
        "labels": "Sprint 2;User Story 4",
        "desc": """User Story 4 — Baseline EEG Emotion Classification
User Story
As a classifier, I want to learn from real EEG feature data
so that baseline emotion recognition performance can be established.
Description
This user story implements and evaluates a baseline EEG emotion classification model without any data augmentation.
Outcome
Baseline performance metrics for EEG-based emotion recognition.
Acceptance Criteria
Classifier model is implemented
Model trains without errors
Baseline accuracy and F1-score are computed
Results are logged and saved"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 3 — GAN-Based EEG Data Augmentation
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "hFjxcB-jdE-KH9Nb7LiIGMkAE3MH",
        "name": "SPRINT 3 — GAN-Based EEG Data Augmentation",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "01/22/2026", "due": "02/01/2026",
        "late": "true",
        "checklist": "Design conditional WGAN-GP architecture;Implement class-conditioned generator and discriminator;Train GAN on EEG feature vectors;Monitor training stability and mode collapse",
        "labels": "Sprint 3;User Story 5",
        "desc": """User Story 5 — GAN Architecture Design & Training
User Story
As a system, I want to generate synthetic EEG features using a GAN model
so that data scarcity and class imbalance are reduced.
Description
This user story designs and trains a conditional Wasserstein GAN with gradient penalty (WGAN-GP) to generate class-conditioned synthetic EEG feature vectors for data augmentation.
Outcome
A trained GAN capable of producing realistic synthetic EEG features.
Acceptance Criteria
GAN architecture is implemented
Training process is stable
Synthetic features match real feature dimensions
No severe mode collapse observed"""
    },
    {
        "id": "WRodLAcQFkKJDCRk3xn29ckAFpjb",
        "name": "SPRINT 3 — GAN-Based EEG Data Augmentation",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "02/02/2026", "due": "02/09/2026",
        "late": "true",
        "checklist": "Validate synthetic feature distributions;Apply static GAN augmentation to training set;Train classifier with augmented data;Record and compare performance metrics",
        "labels": "Sprint 3;User Story 6",
        "desc": """User Story 6 — Synthetic Data Validation & Static Augmentation
User Story
As a researcher, I want to validate synthetic EEG features and apply static augmentation
so that the effectiveness of GAN-generated data can be evaluated.
Description
This user story validates GAN-generated features statistically and evaluates classifier performance when trained with synthetically augmented data.
Outcome
Validated synthetic EEG features and static GAN augmentation results.
Acceptance Criteria
Synthetic samples are validated statistically
Static augmentation is applied correctly
Model performance with GAN data is recorded
Results are reproducible"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 4 — Contrastive EEG Pretraining
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "kR7nXp3QvUa8bWcLmD4f5skAMn2Y",
        "name": "SPRINT 4 — Contrastive EEG Pretraining",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "02/06/2026", "due": "02/14/2026",
        "late": "true",
        "checklist": "Implement NT-Xent contrastive loss function;Design EEG augmentation pipeline (noise, scaling, channel dropout);Build projection head for contrastive learning;Train SimCLR on unlabeled EEG data",
        "labels": "Sprint 4;User Story 7",
        "desc": """User Story 7 — SimCLR Contrastive Loss & EEG Augmentations
User Story
As an EEG encoder, I want to learn robust representations through self-supervised contrastive learning
so that downstream emotion classification benefits from better feature representations.
Description
This user story implements the NT-Xent contrastive loss and designs an EEG-specific augmentation pipeline. Two augmented views of the same EEG sample are pulled together while different samples are pushed apart in embedding space.
Outcome
A contrastive learning framework with EEG-specific augmentations ready for pretraining.
Acceptance Criteria
NT-Xent contrastive loss is implemented correctly
EEG augmentation pipeline includes noise injection, temporal scaling, and channel dropout
Projection head maps encoder outputs to contrastive space
Loss converges during pretraining"""
    },
    {
        "id": "pL9mYr6TwXc2eZfNqG8h1skAJk5W",
        "name": "SPRINT 4 — Contrastive EEG Pretraining",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "02/15/2026", "due": "02/22/2026",
        "late": "true",
        "checklist": "Run full contrastive pretraining on DEAP;Evaluate encoder quality via linear probe;Compare pretrained vs random initialization;Save pretrained encoder checkpoints",
        "labels": "Sprint 4;User Story 8",
        "desc": """User Story 8 — Contrastive Pretraining Execution & Validation
User Story
As a researcher, I want to validate that contrastive pretraining produces better EEG representations
so that downstream tasks benefit from learned features.
Description
This user story executes SimCLR pretraining on the full DEAP dataset and validates the resulting encoder by measuring downstream emotion classification improvement compared to random initialization.
Outcome
Pretrained EEG encoder checkpoints with validated representation quality.
Acceptance Criteria
Contrastive pretraining runs to completion
Linear probe accuracy exceeds random initialization baseline
Pretrained encoder checkpoint is saved
Representation quality is documented"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 5 — Domain Adversarial Adaptation
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "vN2kBs8UyRd5gXhOtJ3i7skAFn9Q",
        "name": "SPRINT 5 — Domain Adversarial Adaptation",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "02/19/2026", "due": "02/27/2026",
        "late": "true",
        "checklist": "Implement gradient reversal layer;Design domain classifier head;Define emotion classification head;Implement lambda scheduling for domain loss",
        "labels": "Sprint 5;User Story 9",
        "desc": """User Story 9 — DANN Architecture & Gradient Reversal Design
User Story
As a multimodal system, I want a gradient reversal mechanism for domain-adversarial training
so that EEG and speech encoders produce domain-invariant representations.
Description
This user story designs the Domain-Adversarial Neural Network (DANN) components including the gradient reversal layer, domain classifier, emotion head, and lambda scheduling for balancing domain and task losses.
Outcome
DANN architecture components ready for cross-domain alignment training.
Acceptance Criteria
Gradient reversal layer is implemented correctly
Domain classifier head distinguishes EEG vs speech domains
Emotion classification head predicts emotion labels
Lambda scheduling balances domain and emotion losses"""
    },
    {
        "id": "wQ4lCs9VzSe6hYiPuK5j8skAHp1R",
        "name": "SPRINT 5 — Domain Adversarial Adaptation",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "02/28/2026", "due": "03/05/2026",
        "late": "false",
        "checklist": "Train DANN on EEG (DEAP) and speech (IEMOCAP);Validate domain-invariant embeddings;Measure domain classification confusion;Save domain-adapted encoder weights",
        "labels": "Sprint 5;User Story 10",
        "desc": """User Story 10 — DANN Cross-Domain Alignment Training
User Story
As a domain adapter, I want to align EEG and speech feature spaces through adversarial training
so that downstream fusion can leverage domain-invariant representations.
Description
This user story trains the DANN model on both EEG (DEAP) and speech (IEMOCAP) data, using adversarial training to make the encoder produce embeddings that are indistinguishable by domain while retaining emotion-discriminative information.
Outcome
Domain-aligned encoder that produces compatible EEG and speech embeddings.
Acceptance Criteria
DANN training runs on both modalities
Domain classifier confusion increases (alignment improves)
Emotion accuracy is maintained or improved
Adapted encoder weights are saved"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 6 — Speech Emotion Processing
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "xS6nDt1WARf7iZjQuL7k9skABr3S",
        "name": "SPRINT 6 — Speech Emotion Processing",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "03/02/2026", "due": "03/09/2026",
        "late": "false",
        "checklist": "Extract MFCC features from IEMOCAP;Implement speech preprocessing pipeline;Design CNN-LSTM speech encoder;Train speech emotion model on IEMOCAP",
        "labels": "Sprint 6;User Story 11",
        "desc": """User Story 11 — Speech Feature Extraction & Encoder Training
User Story
As a speech encoder, I want to process IEMOCAP audio and learn emotion-discriminative features
so that speech-based emotion cues complement EEG-based predictions.
Description
This user story extracts MFCC features from IEMOCAP recordings, designs a CNN-LSTM speech encoder, and trains a speech-only emotion classification model as the speech branch of the multimodal system.
Outcome
A trained speech encoder with extracted IEMOCAP features.
Acceptance Criteria
MFCC features are extracted from IEMOCAP correctly
Speech preprocessing pipeline handles variable-length audio
CNN-LSTM speech encoder is implemented
Speech emotion model trains and produces predictions"""
    },
    {
        "id": "yT8oEu2XBSg8jAkRvM9l0skACt5T",
        "name": "SPRINT 6 — Speech Emotion Processing",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "03/10/2026", "due": "03/15/2026",
        "late": "false",
        "checklist": "Implement contrastive speech pretraining;Apply label-aligned EEG-speech pairing;Validate cross-modal pairing quality;Compute speech-only baseline metrics",
        "labels": "Sprint 6;User Story 12",
        "desc": """User Story 12 — Speech Pretraining & Cross-Modal Pairing
User Story
As a multimodal system, I want speech representations pretrained and label-aligned EEG-speech pairs created
so that cross-modal fusion receives semantically consistent inputs.
Description
This user story implements contrastive pretraining for the speech encoder and establishes label-aligned pairing between EEG (DEAP) and speech (IEMOCAP) samples based on emotion class rather than index, ensuring fusion receives consistent supervision signals.
Outcome
Pretrained speech encoder and label-aligned cross-modal pairs.
Acceptance Criteria
Contrastive speech pretraining is implemented
EEG-speech pairs are aligned by emotion label (not index)
Pairing quality is validated
Speech-only baseline accuracy is recorded"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 7 — Cross-Modal Fusion Architecture
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "EN8NueYXgUC9ebzyIs_WIskALAGR",
        "name": "SPRINT 7 — Cross-Modal Fusion Architecture",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "03/12/2026", "due": "03/19/2026",
        "late": "false",
        "checklist": "Design cross-modal multi-head attention module;Implement gated fusion mechanism;Define focal loss with class weighting;Integrate attention module with pretrained encoders",
        "labels": "Sprint 7;User Story 13",
        "desc": """User Story 13 — CMMA Cross-Modal Multi-Head Attention Design
User Story
As a multimodal classifier, I want to jointly attend to EEG and speech features via cross-modal attention
so that inter-modal emotion correlations are captured for better classification.
Description
This user story designs and implements the Cross-Modal Multi-Head Attention (CMMA) fusion module. EEG attends to speech and speech attends to EEG through multi-head cross-attention, followed by gated fusion and an emotion classification head. Focal loss with class weighting addresses DEAP's 37x class imbalance.
Outcome
A cross-modal attention fusion architecture integrated with pretrained encoders.
Acceptance Criteria
Cross-modal multi-head attention architecture is designed
Gated fusion mechanism combines attended features
Focal loss with sqrt class weighting handles imbalance
Module integrates with pretrained EEG and speech encoders"""
    },
    {
        "id": "FYN2et-LYkC04FMtb2Sj2ckAGSvh",
        "name": "SPRINT 7 — Cross-Modal Fusion Architecture",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "03/20/2026", "due": "03/26/2026",
        "late": "false",
        "checklist": "Implement transformer fusion baseline;Compare CMMA vs concatenation vs transformer fusion;Benchmark fusion strategies on validation set;Select best fusion architecture",
        "labels": "Sprint 7;User Story 14",
        "desc": """User Story 14 — Fusion Strategy Comparison & Selection
User Story
As a researcher, I want to compare multiple fusion strategies
so that the best-performing architecture is selected for the final pipeline.
Description
This user story implements and benchmarks multiple fusion approaches: simple concatenation (v1), transformer attention (v3), and CMMA cross-attention (v5). Performance is compared on the validation set to select the final architecture.
Outcome
Empirical comparison of fusion strategies with CMMA selected as best.
Acceptance Criteria
Concatenation, transformer, and CMMA fusion are all implemented
Each fusion strategy is evaluated on the same validation set
Performance comparison is documented
Best fusion architecture is selected"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 8 — End-to-End Pipeline Training & Optimization
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "aU0pFv3YCSh9kBlSwN1m2skADv7U",
        "name": "SPRINT 8 — End-to-End Pipeline Training & Optimization",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "03/23/2026", "due": "03/30/2026",
        "late": "false",
        "checklist": "Train end-to-end CMMA pipeline (encoders + fusion + classifier);Implement learning rate scheduling and early stopping;Run iterative architecture versions (v1 through v5);Log performance metrics across all versions",
        "labels": "Sprint 8;User Story 15",
        "desc": """User Story 15 — End-to-End CMMA Training & Version Iteration
User Story
As a system, I want to train the complete CMMA pipeline end-to-end with iterative optimization
so that multimodal emotion recognition accuracy is maximized.
Description
This user story trains the full pipeline (pretrained EEG encoder + pretrained speech encoder + CMMA fusion + classifier) end-to-end. Multiple architecture versions (v1 through v5.3) are trained iteratively, with each version addressing limitations discovered in the previous one.
Outcome
An optimized multimodal emotion recognition system with documented version progression.
Acceptance Criteria
End-to-end CMMA pipeline trains without errors
Learning rate scheduling and early stopping are applied
Multiple versions (v1-v5.3) are trained and compared
Best model achieves target accuracy and F1-score"""
    },
    {
        "id": "bV1qGw4ZDTi0lCmTxO3n3skAEw8V",
        "name": "SPRINT 8 — End-to-End Pipeline Training & Optimization",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "03/31/2026", "due": "04/05/2026",
        "late": "false",
        "checklist": "Perform hyperparameter search (learning rate, dropout, attention heads);Tune focal loss gamma and class weight strategy;Optimize batch size and pairing strategy;Record best hyperparameter configuration",
        "labels": "Sprint 8;User Story 16",
        "desc": """User Story 16 — Hyperparameter Tuning & Configuration Optimization
User Story
As a researcher, I want to systematically tune hyperparameters
so that the model achieves peak performance under the best configuration.
Description
This user story performs systematic hyperparameter tuning for the CMMA pipeline including learning rate, dropout rate, number of attention heads, focal loss gamma, class weight strategy (sqrt vs inverse), batch size, and EEG-speech pairing count.
Outcome
Optimal hyperparameter configuration for the CMMA pipeline.
Acceptance Criteria
Hyperparameter search covers learning rate, dropout, and attention heads
Focal loss gamma and class weight strategy are tuned
Batch size and pairing strategy are optimized
Best configuration is documented and reproducible"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 9 — LOSO Subject-Independent Evaluation
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "naN5yJwPv0SxicdesFQFYMkAGkH4",
        "name": "SPRINT 9 — LOSO Subject-Independent Evaluation",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "04/02/2026", "due": "04/10/2026",
        "late": "false",
        "checklist": "Implement 32-fold LOSO evaluation framework;Run LOSO v1 baseline with global checkpoints;Implement improved LOSO v2 with cross-subject normalization;Compare LOSO v1 and v2 results",
        "labels": "Sprint 9;User Story 17",
        "desc": """User Story 17 — LOSO v1 & v2 Implementation
User Story
As a researcher, I want to evaluate the system using Leave-One-Subject-Out cross-validation
so that reported performance reflects true generalization to unseen subjects.
Description
This user story implements 32-fold LOSO evaluation where each DEAP subject is held out as the test set. LOSO v1 establishes a baseline, while LOSO v2 adds cross-subject normalization, improved class weighting, multi-pairing ensemble, and optimized training for improved subject-independent accuracy.
Outcome
Subject-independent accuracy with confidence intervals across 32 folds.
Acceptance Criteria
32-fold LOSO framework is implemented
LOSO v1 baseline results are recorded
LOSO v2 with improvements is implemented
Per-fold results include mean, std, and per-class metrics"""
    },
    {
        "id": "cW2rHx5AETj1mDnUyP5o4skAFx9W",
        "name": "SPRINT 9 — LOSO Subject-Independent Evaluation",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "04/11/2026", "due": "04/17/2026",
        "late": "false",
        "checklist": "Identify encoder pretraining data leak in LOSO v2;Implement per-fold contrastive pretraining (train subjects only);Implement per-fold DANN training (no global checkpoint);Run strict 32-fold LOSO and compare with leaked results",
        "labels": "Sprint 9;User Story 18",
        "desc": """User Story 18 — Strict Leak-Free LOSO Evaluation
User Story
As a researcher, I want a fully leak-free subject-independent evaluation
so that reported accuracy reflects true generalization without any data contamination.
Description
This user story identifies the encoder pretraining data leak (contrastive + DANN used all 32 subjects) and implements Strict LOSO v2 where ALL pretraining is moved inside the fold loop, ensuring the held-out test subject's data never influences encoder weights.
Outcome
Unbiased subject-independent accuracy with zero data leakage.
Acceptance Criteria
Encoder pretraining leak is identified and documented
Per-fold contrastive pretraining uses only 30 train subjects
Per-fold DANN training uses only train subjects
Leak impact is quantified by comparing strict vs standard LOSO"""
    },

    # ═══════════════════════════════════════════════════════════════
    # SPRINT 10 — Final Evaluation, Ablation & Reporting
    # ═══════════════════════════════════════════════════════════════
    {
        "id": "Xs0d0YMKOEaa1YGINc1wxMkAIp56",
        "name": "SPRINT 10 — Final Evaluation, Ablation & Reporting",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "04/14/2026", "due": "04/21/2026",
        "late": "false",
        "checklist": "Perform ablation experiments (remove contrastive, DANN, CMMA individually);Compute accuracy, F1-score, UAR, and Cohen's kappa;Generate confusion matrices and per-class metrics;Create result comparison tables across all versions",
        "labels": "Sprint 10;User Story 19",
        "desc": """User Story 19 — Ablation Studies & Comprehensive Metrics
User Story
As a researcher, I want to perform ablation experiments and compute comprehensive metrics
so that the contribution of each component is quantified.
Description
This user story systematically removes individual components (contrastive pretraining, DANN, CMMA, focal loss) and measures the impact on performance. Full metrics including accuracy, macro F1, UAR, and Cohen's kappa are computed for all model versions.
Outcome
Ablation results quantifying each component's contribution.
Acceptance Criteria
Ablation removes each component individually
All metrics are computed for each ablation variant
Results are compared in structured tables
Component contributions are clearly quantified"""
    },
    {
        "id": "dX3sIy6BFUk2nEoVzQ7p5skAGy0X",
        "name": "SPRINT 10 — Final Evaluation, Ablation & Reporting",
        "bucket": "BACKLOG",
        "progress": "Not started",
        "start": "04/22/2026", "due": "04/28/2026",
        "late": "false",
        "checklist": "Write LaTeX paper with all results;Generate publication-quality plots and figures;Finalize VERSION_HISTORY documentation;Prepare reproducibility package",
        "labels": "Sprint 10;User Story 20",
        "desc": """User Story 20 — Documentation, Paper Writing & Reproducibility
User Story
As a researcher, I want to document all results and prepare a publication-ready paper
so that findings are reproducible and suitable for academic submission.
Description
This user story finalizes all documentation including the LaTeX research paper, result visualizations, version history, and a reproducibility package that allows others to replicate the experiments.
Outcome
Complete documentation and publication-ready paper.
Acceptance Criteria
LaTeX paper includes abstract, methodology, results, and discussion
Plots and figures are publication-quality
VERSION_HISTORY documents the full v1 to v5.3 progression
Experimental setup is fully reproducible"""
    },
]

# ── Write headers ──
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Sprint colors (cycle through 5 colors for 10 sprints) ──
sprint_fills = {
    "SPRINT 1": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "SPRINT 2": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "SPRINT 3": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "SPRINT 4": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "SPRINT 5": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    "SPRINT 6": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "SPRINT 7": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
    "SPRINT 8": PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid"),
    "SPRINT 9": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
    "SPRINT 10": PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),
}

# ── Write data rows ──
for i, row in enumerate(rows, 2):
    sprint_key = row["name"].split(" —")[0]
    fill = sprint_fills.get(sprint_key, PatternFill())

    data = [
        row["id"],
        row["name"],
        row["bucket"],
        row["progress"],
        "Medium",
        ASSIGNED,
        CREATED_BY,
        CREATED_DATE,
        row["start"],
        row["due"],
        "false",
        row["late"],
        "",  # Completed Date
        row["checklist"],
        row["labels"],
        row["desc"],
    ]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# ── Column widths ──
widths = {
    "A": 32, "B": 52, "C": 15, "D": 12, "E": 10, "F": 55, "G": 45,
    "H": 14, "I": 14, "J": 14, "K": 12, "L": 8, "M": 16,
    "N": 70, "O": 25, "P": 85
}
for col_letter, width in widths.items():
    ws.column_dimensions[col_letter].width = width

# ── Save ──
out_path = r"c:\Users\ravin\Downloads\MAJORPROJECT\amers\docs\AMERS_Backlog_10sprints.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
